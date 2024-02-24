from tkinter import *
import os
import sys
from tkinter import filedialog
import openpyxl as opx

wdir = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(wdir)


class Contact:
    def __init__(self, user_data):
        self.l_name = user_data[0]
        self.f_name = user_data[1]
        self.m_name = user_data[2]
        self.phone_number = user_data[3]
        self.birthday = user_data[4]
        self.email = user_data[5]
        self.comment = user_data[6]
        self.image_url = user_data[7] if len(user_data) > 6 else ''
        self.relationship = user_data[8] if len(user_data) > 7 else ''

    def toStr(self):
        contact_str = self.f_name
        if self.l_name:
            contact_str += ' ' + self.l_name[0] + '.'
        contact_str += ' ' + self.phone_number
        if self.comment:
            contact_str += ' ' + self.comment
        return contact_str

    def toRow(self):
        return [self.l_name, self.f_name, self.m_name, self.phone_number, self.birthday, self.email, self.comment,
                self.image_url, self.relationship]

    def toList(self):
        return [self.l_name, self.f_name, self.m_name, self.phone_number, self.birthday, self.email, self.comment,
                self.image_url, self.relationship]


class Main(Frame):
    config = {'save_to': 'xlsx', 'save_file_name': 'contacts'}
    contacts = []

    def __init__(self, master):
        super().__init__(master)
        self.init_main()

    def init_main(self):
        toolbar = Frame(bg="#d7d8e0")
        toolbar.pack(side=TOP, fill=X)

        Button(toolbar, text="Добавить", command=self.add_contact).pack(side=LEFT)
        Button(toolbar, text="Изменить").pack(side=LEFT)

        Button(toolbar, text="Сохранить", command=self.saveToFile).pack(side=RIGHT)
        Button(toolbar, text="Открыть", command=self.openFromFile).pack(side=RIGHT)

        mainframe = Frame(bg="#e8e8e8")
        mainframe.pack(side=TOP, fill=BOTH)

        scrollbar = Scrollbar(mainframe)
        scrollbar.pack(side=RIGHT, fill=Y)

        self.listbox = Listbox(mainframe, width=100, height=480, yscrollcommand=scrollbar.set)
        self.listbox.pack(side=LEFT, fill=BOTH)
        scrollbar.config(command=self.listbox.yview)
        self.openFromFile()

    def toValue(self) -> list:
        result_list = []
        for contact in self.contacts:
            result_list.append(contact.toStr())
        return result_list

    def saveToFile(self):
        if self.config['save_to'] != 'xlsx':
            return

        xlsx = opx.Workbook()
        active_sheet = xlsx.active

        for index, contact in enumerate(self.contacts):
            active_sheet.append(contact.toList())

        file_name = f"{self.config['save_file_name']}.{self.config['save_to']}"
        xlsx.save(file_name)

    def openFromFile(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path:
            return

        xls = opx.load_workbook(file_path)
        active_sheet = xls.active

        self.contacts.clear()
        for row in active_sheet.iter_rows(values_only=True):
            self.contacts.append(Contact(row))
        self.reRenderListBox()

    def reRenderListBox(self):
        self.listbox.delete(0, END)
        for contact in self.contacts:
            self.listbox.insert(END, contact.toStr())

    def add_contact(self):
        new_contact = Contact(["", "", "", "", "", "", "", "", ""])
        self.contacts.append(new_contact)
        ContactEditFrame(new_contact, self)
        self.reRenderListBox()


class ContactEditFrame:
    def __init__(self, contact, main):
        self.contact = contact
        self.main = main
        self.init_main()

    def init_main(self):
        self.edit_window = Tk()
        self.edit_window.title('Добавить контакт')

        Label(self.edit_window, text='Фамилия:').grid(row=0, column=0)
        self.l_name_entry = Entry(self.edit_window)
        self.l_name_entry.grid(row=0, column=1)

        Label(self.edit_window, text='Имя:').grid(row=1, column=0)
        self.f_name_entry = Entry(self.edit_window)
        self.f_name_entry.grid(row=1, column=1)

        Label(self.edit_window, text='Номер телефона:').grid(row=2, column=0)
        self.phone_entry = Entry(self.edit_window)
        self.phone_entry.grid(row=2, column=1)

        Label(self.edit_window, text='Дата рождения:').grid(row=3, column=0)
        self.birthday_entry = Entry(self.edit_window)
        self.birthday_entry.grid(row=3, column=1)

        Label(self.edit_window, text='Email:').grid(row=4, column=0)
        self.email_entry = Entry(self.edit_window)
        self.email_entry.grid(row=4, column=1)

        Label(self.edit_window, text='Комментарий:').grid(row=5, column=0)
        self.comment_entry = Entry(self.edit_window)
        self.comment_entry.grid(row=5, column=1)

        Label(self.edit_window, text='Отношения:').grid(row=6, column=0)
        self.relationship_entry = Entry(self.edit_window)
        self.relationship_entry.grid(row=6, column=1)

        Label(self.edit_window, text='Изображение:').grid(row=7, column=0)
        self.image_path_entry = Entry(self.edit_window)
        self.image_path_entry.grid(row=7, column=1)
        image_button = Button(self.edit_window, text='Выбрать изображение', command=self.choose_image)
        image_button.grid(row=7, column=2)

        Button(self.edit_window, text='Сохранить', command=self.save_contact).grid(row=8, column=0, columnspan=2)

    def choose_image(self):
        file_path = filedialog.askopenfilename(title="Выберите изображение",
                                               filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
        self.image_path_entry.delete(0, END)
        self.image_path_entry.insert(END, file_path)

    def save_contact(self):
        l_name = self.l_name_entry.get()
        f_name = self.f_name_entry.get()
        phone = self.phone_entry.get()
        birthday = self.birthday_entry.get()
        email = self.email_entry.get()
        comment = self.comment_entry.get()
        relationship = self.relationship_entry.get()
        image_path = self.image_path_entry.get()

        self.contact.l_name = l_name
        self.contact.f_name = f_name
        self.contact.phone_number = phone
        self.contact.birthday = birthday
        self.contact.email = email
        self.contact.comment = comment
        self.contact.relationship = relationship
        self.contact.image_url = image_path

        self.edit_window.destroy()
        self.main.reRenderListBox()


if __name__ == "__main__":
    window = Tk()
    Main(window).pack()
    window.title("Контакты")
    window.geometry("720x480")
    window.resizable(False, False)
    window.mainloop()


